"""A helper script to parse my old budget file to generic csv."""
import argparse
import csv
import logging
from typing import List, Optional

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.comments import Comment
from openpyxl.worksheet.worksheet import Worksheet
from python_scripts.execution_time import log_execution_time
from python_scripts.expenditure import Expenditure

logging.basicConfig(level=logging.INFO)

ROWS = range(2, 32 + 1)
COLUMN_DATE = 1
COLUMNS = {
    1: "DATE",
    3: "bills-flat",
    4: "bills-other",
    6: "food-shop",
    7: "food-restaurant",
    8: "food-sweets",
    9: "food-alcohol",
    10: "cosmetics",
    11: "medicines",
    12: "domestic",
    13: "clothes",
    15: "party",
    16: "sport",
    17: "school",
    18: "other",
    19: "other-mom",
}


def get_filename_from_args() -> str:  # noqa: D103
    parser = argparse.ArgumentParser(description="Process some budget files.")
    parser.add_argument(
        "filename", metavar="F", type=str, help="Budget filename to be parsed"
    )
    args = parser.parse_args()
    return args.filename


def get_worksheets(file: str) -> List[Worksheet]:  # noqa: D103
    wb = load_workbook(file, data_only=True)
    logging.info('Opened workbook: "%s"', file)
    return [ws for ws in wb.worksheets if ws.title != "BUDGET"]


def sanitize_comment(comment: Comment) -> Optional[str]:  # noqa: D103
    if comment is None:
        return None
    ret_val = comment.content.strip()
    ret_val = ret_val.strip("\\n")
    ret_val = ret_val.strip("\n")
    ret_val = ret_val.strip(":")
    return ret_val


def parse_cell(
    cell: Cell, row_nr: int, column_nr: int, worksheet: Worksheet
) -> Expenditure:  # noqa: D103
    date = worksheet.cell(row=row_nr, column=COLUMN_DATE).value
    return Expenditure(
        value=float(cell.value),
        date=date,
        category=COLUMNS[column_nr],
        comment=sanitize_comment(cell.comment),
    )


def fetch_expendeitures(worksheets: List[Worksheet]) -> List[Expenditure]:  # noqa: D103
    expenditures = list()
    for worksheet in worksheets:
        for row_nr in ROWS:
            for column_nr in COLUMNS:
                if column_nr == COLUMN_DATE:  # skip date column
                    continue

                cell = worksheet.cell(row=row_nr, column=column_nr)
                if cell.value:
                    parsed_cell = parse_cell(cell, row_nr, column_nr, worksheet)
                    expenditures.append(parsed_cell)
    logging.info("Fetched %d expenditures", len(expenditures))
    return sorted(expenditures, key=lambda x: x.date)


def save_expenditures_to_csv(
    expenditures: List[Expenditure], file
) -> None:  # noqa: D103
    with open(file, mode="w") as f:
        writer = csv.writer(f, delimiter=",")
        for e in expenditures:
            writer.writerow([e.date, e.value, e.category, e.comment])
    logging.info('Saved file: "%s"', file)


@log_execution_time
def main(file=file):
    worksheets = get_worksheets(file)
    expenditures = fetch_expendeitures(worksheets)
    save_expenditures_to_csv(expenditures, "expenditures.csv")


if __name__ == "__main__":
    file = get_filename_from_args()
    main(file)
